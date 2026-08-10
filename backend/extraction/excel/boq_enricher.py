import os
from typing import Any

import openpyxl

from backend.extraction.excel.excel_tool import resolve_document_path
from backend.storage.postgres_store import PostgresStore


class BOQPriceEnricherTool:
    name = "enrich_boq_excel"
    description = (
        "Automatically enrich an unpriced BOQ Excel file with prices from the DB price catalog.\n"
        "RULES:\n"
        "1. Provide `filename` of the BOQ (e.g., 'Fiber BOQ.xlsx').\n"
        "2. Provide `part_number_col` (the column letter containing part numbers, e.g., 'A').\n"
        "3. Provide `qty_col` (the column letter containing quantities, e.g., 'B').\n"
        "4. This tool outputs an enriched Excel file path and a summary.\n"
    )
    from typing import ClassVar
    input_schema: ClassVar[dict] = {
        "type": "object",
        "properties": {
            "filename": {"type": "string"},
            "part_number_col": {"type": "string", "default": "A"},
            "qty_col": {"type": "string", "default": "B"}
        },
        "required": ["filename"]
    }

    def run(self, **kwargs) -> dict[str, Any]:
        filename = kwargs.get("filename")
        part_number_col = kwargs.get("part_number_col", "A")
        qty_col = kwargs.get("qty_col", "B")
        
        if not filename:
            return {"error": "filename is required."}
            
        try:
            path = resolve_document_path(filename)
            wb = openpyxl.load_workbook(path)
            ws = wb.active

            # Find the next available columns for Unit Price and Total Price
            max_col = ws.max_column
            unit_price_col_idx = max_col + 1
            total_price_col_idx = max_col + 2
            
            unit_price_col_letter = openpyxl.utils.get_column_letter(unit_price_col_idx)
            total_price_col_letter = openpyxl.utils.get_column_letter(total_price_col_idx)

            # Add headers
            ws[f"{unit_price_col_letter}1"] = "Unit Price"
            ws[f"{total_price_col_letter}1"] = "Total Price"

            # Connect to DB and fetch price catalog
            # We fetch all prices into a dictionary to avoid querying per row
            store = PostgresStore()
            rows = store.conn.execute("SELECT part_number, unit_price FROM price_catalog").fetchall()
            price_map = {str(r[0]).strip(): float(r[1]) for r in rows}

            enriched_count = 0
            
            # Start from row 2 assuming row 1 is header
            for row in range(2, ws.max_row + 1):
                part_number_cell = ws[f"{part_number_col}{row}"].value
                if part_number_cell:
                    part_str = str(part_number_cell).strip()
                    if part_str in price_map:
                        unit_price = price_map[part_str]
                        # Inject Unit Price
                        ws[f"{unit_price_col_letter}{row}"] = unit_price
                        
                        # Inject Formula: =Qty * UnitPrice
                        ws[f"{total_price_col_letter}{row}"] = f"={qty_col}{row}*{unit_price_col_letter}{row}"
                        enriched_count += 1
            
            # Save the enriched file
            out_filename = filename.replace(".xlsx", "_Enriched.xlsx")
            out_path = os.path.join(os.path.dirname(path), out_filename)
            wb.save(out_path)
            
            return {
                "blocks": [{"type": "text", "content": f"Successfully enriched {enriched_count} items. Saved to {out_filename}."}],
                "file_path": out_path,
                "enriched_count": enriched_count
            }
        except Exception as e:
            return {"error": f"BOQ Enrichment failed: {e!s}"}
