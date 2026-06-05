import os
import uuid
import json
import pandas as pd
from typing import List, Dict, Any, Optional


class ExcelExtractorTool:
    """
    Extracts tables, embedded images, formulas, and pivot table data from Excel files.
    Output strictly follows the NormalizedBlock contract in schemas.py.
    """

    def run(self, state: Dict[str, Any], config: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Entry point for the pipeline.

        state  — expects state["file_path"] and optionally state["document_id"]
        config — pipeline config dict (extraction_confidence, default_language, etc.)
        """
        file_path   = state["file_path"]
        document_id = state.get("document_id")
        return self._extract(file_path, document_id=document_id, config=config)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _extract(
        self,
        file_path: str,
        document_id: Optional[str] = None,
        config: Optional[Dict[str, Any]] = None,
    ) -> List[Dict[str, Any]]:
        blocks: List[Dict[str, Any]] = []
        cfg      = config or {}
        doc_id   = str(document_id) if document_id else str(uuid.uuid4())
        filename = os.path.basename(file_path)

        try:
            wb = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
        except Exception as e:
            print(f"[excel_extractor] Failed to open {file_path}: {e}")
            return blocks

        # ── 1. Table blocks ────────────────────────────────────────────
        for sheet_name, df in wb.items():
            try:
                df = df.dropna(how="all").dropna(axis=1, how="all")
                if df.empty:
                    continue
                df = df.fillna("")

                block_id   = str(uuid.uuid4())
                cell_range = f"{sheet_name}!A1:{self._col_letter(len(df.columns))}{len(df) + 1}"

                block = {
                    "block_id":    block_id,
                    "document_id": doc_id,
                    "type":        "table",
                    # df.to_markdown() — this is what gets embedded and searched
                    "text": df.to_markdown(index=False),
                    "table_data": {
                        "headers": [str(c) for c in df.columns.tolist()],
                        "rows":    df.values.tolist(),
                    },
                    "source_ref": {
                        "filename": filename,
                        "page":     None,
                        "sheet":    str(sheet_name),
                        "slide":    None,
                        "bbox":     None,
                    },
                    "confidence": cfg.get("extraction_confidence", 1.0),
                    "language":   cfg.get("default_language", "en"),
                    "metadata": {
                        # cell_range stays here — does NOT travel downstream
                        "cell_range":        cell_range,
                        "enrichment_failed": cfg.get("enrichment_failed_flag", False),
                    },
                }
                blocks.append(block)

            except Exception as e:
                print(f"[excel_extractor] Skipping sheet '{sheet_name}': {e}")
                continue

        # ── 2. Embedded image blocks ───────────────────────────────────
        image_blocks = self._extract_images(file_path, doc_id, filename, cfg)
        blocks.extend(image_blocks)

        # ── 3. Formula blocks ──────────────────────────────────────────
        formula_blocks = self._extract_formulas(file_path, doc_id, filename, cfg)
        blocks.extend(formula_blocks)

        # ── 4. Pivot table blocks ──────────────────────────────────────
        pivot_blocks = self._extract_pivots(file_path, doc_id, filename, cfg)
        blocks.extend(pivot_blocks)

        return blocks

    def _extract_images(
        self,
        file_path: str,
        doc_id: str,
        filename: str,
        cfg: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        """
        Pulls embedded images out of the workbook via openpyxl directly.
        Saves raw bytes to uploads/images/<doc_id>/<block_id>_raw.png.
        Returns image blocks with metadata["raw_image_path"] set.
        vision_enrichment_tool picks them up from there.
        """
        import openpyxl

        blocks: List[Dict[str, Any]] = []

        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception as e:
            print(f"[excel_extractor] openpyxl open failed for images: {e}")
            return blocks

        out_dir = os.path.join("uploads", "images", doc_id)
        os.makedirs(out_dir, exist_ok=True)

        for sheet in wb.worksheets:
            for image in getattr(sheet, "_images", []):
                try:
                    block_id = str(uuid.uuid4())
                    raw_path = os.path.join(out_dir, f"{block_id}_raw.png")

                    # image.ref is a BytesIO — write raw bytes to disk
                    image_data = image.ref.read() if hasattr(image.ref, "read") else bytes(image.ref)
                    with open(raw_path, "wb") as f:
                        f.write(image_data)

                    block = {
                        "block_id":    block_id,
                        "document_id": doc_id,
                        "type":        "image",
                        # text is empty — vision_enrichment_tool fills this in
                        "text": None,
                        "table_data":  None,
                        "source_ref": {
                            "filename": filename,
                            "page":     None,
                            "sheet":    sheet.title,
                            "slide":    None,
                            "bbox":     None,
                        },
                        "confidence": cfg.get("extraction_confidence", 1.0),
                        "language":   cfg.get("default_language", "en"),
                        "metadata": {
                            # extractor sets raw_image_path; vision sets image_path + text
                            "raw_image_path":   raw_path,
                            "enrichment_failed": False,
                        },
                    }
                    blocks.append(block)

                except Exception as e:
                    print(f"[excel_extractor] Skipping image in sheet '{sheet.title}': {e}")
                    continue

        return blocks

    def _extract_formulas(
        self,
        file_path: str,
        doc_id: str,
        filename: str,
        cfg: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        """
        Extracts cells containing formulas.
        Emits one block per formula cell — type="formula".
        text = the formula string (searchable + embeddable).
        metadata["cell_ref"] = e.g. "Sheet1!B4"
        metadata["formula"]  = "=SUMIF(A2:A20,\"Q1\",C2:C20)"

        Uses data_only=False so openpyxl returns formula strings, not cached values.
        """
        import openpyxl

        blocks: List[Dict[str, Any]] = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
        except Exception as e:
            print(f"[excel_extractor] Formula load failed: {e}")
            return blocks

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type != "f" or cell.value is None:
                        continue
                    try:
                        formula  = str(cell.value)
                        cell_ref = f"{sheet.title}!{cell.coordinate}"
                        block_id = str(uuid.uuid4())

                        block = {
                            "block_id":    block_id,
                            "document_id": doc_id,
                            "type":        "formula",
                            "text":        formula,
                            "table_data":  None,
                            "source_ref": {
                                "filename": filename,
                                "page":     None,
                                "sheet":    sheet.title,
                                "slide":    None,
                                "bbox":     None,
                            },
                            "confidence": cfg.get("extraction_confidence", 1.0),
                            "language":   cfg.get("default_language", "en"),
                            "metadata": {
                                "cell_ref":          cell_ref,
                                "formula":           formula,
                                "sheet":             sheet.title,
                                "enrichment_failed": False,
                            },
                        }
                        blocks.append(block)

                    except Exception as e:
                        print(f"[excel_extractor] Skipping formula at {cell.coordinate}: {e}")
                        continue

        return blocks

    def _extract_pivots(
        self,
        file_path: str,
        doc_id: str,
        filename: str,
        cfg: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        """
        Extracts pivot table source data from cache records.
        Pivot display values aren't in cells — they're computed by Excel at render time.
        We pull from cacheDefinition.records instead.

        Emits one "table" block per pivot.
        metadata["pivot_name"]         = pivot table name
        metadata["pivot_source_range"] = source ref string e.g. "Sheet1!A1:D200"
        metadata["cell_range"]         = same, for citation (does not travel downstream)
        """
        import openpyxl

        blocks: List[Dict[str, Any]] = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            print(f"[excel_extractor] Pivot load failed: {e}")
            return blocks

        for sheet in wb.worksheets:
            for pivot in getattr(sheet, "_pivots", []):
                try:
                    pivot_name   = getattr(pivot, "name", "unnamed_pivot")
                    cache_def    = getattr(pivot, "cacheDefinition", None)
                    source_range = ""

                    if cache_def:
                        src = getattr(cache_def, "cacheSource", None)
                        if src and hasattr(src, "worksheetSource"):
                            ws_src       = src.worksheetSource
                            source_range = getattr(ws_src, "ref", "")

                    # Pull field names from cacheFields
                    headers = []
                    if cache_def and hasattr(cache_def, "cacheFields"):
                        headers = [
                            getattr(cf, "name", f"field_{i}")
                            for i, cf in enumerate(cache_def.cacheFields)
                        ]

                    # Pull rows from cacheRecords
                    rows = []
                    if cache_def and hasattr(cache_def, "records") and cache_def.records:
                        for record in cache_def.records.r:
                            row = []
                            for item in record:
                                row.append(getattr(item, "v", None))
                            rows.append(row)

                    if not headers and not rows:
                        continue

                    df       = pd.DataFrame(rows, columns=headers if headers else None)
                    block_id = str(uuid.uuid4())

                    block = {
                        "block_id":    block_id,
                        "document_id": doc_id,
                        "type":        "table",
                        "text":        df.to_markdown(index=False),
                        "table_data": {
                            "headers": headers,
                            "rows":    rows,
                        },
                        "source_ref": {
                            "filename": filename,
                            "page":     None,
                            "sheet":    sheet.title,
                            "slide":    None,
                            "bbox":     None,
                        },
                        "confidence": cfg.get("extraction_confidence", 1.0),
                        "language":   cfg.get("default_language", "en"),
                        "metadata": {
                            "pivot_name":         pivot_name,
                            "pivot_source_range": source_range,
                            "cell_range":         source_range,  # for citation, doesn't travel downstream
                            "enrichment_failed":  False,
                        },
                    }
                    blocks.append(block)

                except Exception as e:
                    print(f"[excel_extractor] Skipping pivot '{getattr(pivot, 'name', '?')}': {e}")
                    continue

        return blocks

    @staticmethod
    def _col_letter(n: int) -> str:
        """Convert 1-based column index to Excel letter (1→A, 27→AA)."""
        result = ""
        while n:
            n, r = divmod(n - 1, 26)
            result = chr(65 + r) + result
        return result or "A"


# ------------------------------------------------------------------
# SANDBOX TEST
# ------------------------------------------------------------------
if __name__ == "__main__":
    test_file = "test-data/test.xlsx"

    mock_state  = {"file_path": test_file, "document_id": "doc-001"}
    mock_config = {"extraction_confidence": 0.95, "default_language": "en"}

    tool    = ExcelExtractorTool()
    results = tool.run(mock_state, mock_config)

    tables   = [b for b in results if b["type"] == "table"]
    images   = [b for b in results if b["type"] == "image"]
    formulas = [b for b in results if b["type"] == "formula"]
    pivots   = [b for b in results if b["type"] == "table" and "pivot_name" in b["metadata"]]

    print(f"Extracted {len(tables)} table(s), {len(images)} image(s), {len(formulas)} formula(s), {len(pivots)} pivot(s).\n")

    if tables:
        print("--- First table block ---")
        preview = {k: v for k, v in tables[0].items() if k != "table_data"}
        preview["text_preview"] = (tables[0]["text"] or "")[:300]
        print(json.dumps(preview, indent=2))

    if formulas:
        print("\n--- First formula block ---")
        print(json.dumps(formulas[0], indent=2))

    if images:
        print("\n--- First image block ---")
        print(json.dumps(images[0], indent=2))

    if pivots:
        print("\n--- First pivot block ---")
        preview = {k: v for k, v in pivots[0].items() if k != "table_data"}
        preview["text_preview"] = (pivots[0]["text"] or "")[:300]
        print(json.dumps(preview, indent=2))
