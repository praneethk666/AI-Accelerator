import os
import uuid
from io import BytesIO
from datetime import date, datetime

import pandas as pd
from typing import List, Dict, Any, Optional
from backend.core.paths import display_filename
from backend.core.schemas import NormalizedBlock, SourceRef, as_dicts


def _detect_image_ext(data: bytes) -> str:
    """Sniff an embedded image's real format from its magic bytes (Excel/PPT can
    embed jpg/gif/emf/... — saving them all as .png corrupts the file)."""
    if not data:
        return "png"
    if data[:8] == b"\x89PNG\r\n\x1a\n": return "png"
    if data[:3] == b"\xff\xd8\xff": return "jpg"
    if data[:6] in (b"GIF87a", b"GIF89a"): return "gif"
    if data[:4] == b"RIFF" and data[8:12] == b"WEBP": return "webp"
    if data[:4] in (b"MM\x00*", b"II*\x00"): return "tiff"
    if data[:2] == b"BM": return "bmp"
    if data[:2] == b"\x1f\x8b": return "gz"
    return "png"


def _save_image_blob(image_data: bytes, out_dir: str, block_id: str, *, source_name: str | None = None) -> tuple[str, str]:
    """Persist an embedded image safely and return (raw_path, preview_path).

    We keep the original bytes on disk when possible, but if Pillow can decode
    them we also create a PNG preview for browser/UI consumption.
    """
    os.makedirs(out_dir, exist_ok=True)
    source_name = source_name or ""
    ext = _detect_image_ext(image_data)
    raw_path = os.path.join(out_dir, f"{block_id}_raw.{ext}")
    with open(raw_path, "wb") as f:
        f.write(image_data)

    preview_path = raw_path
    try:
        from PIL import Image

        img = Image.open(BytesIO(image_data))
        img = img.convert("RGBA") if img.mode not in ("RGB", "RGBA") else img
        preview_path = os.path.join(out_dir, f"{block_id}.png")
        img.save(preview_path, format="PNG")
    except Exception:
        # If we can't decode it, keep the raw bytes and let downstream decide.
        preview_path = raw_path

    return raw_path, preview_path


def _json_safe_cell(value):
    """Convert Excel cell values to stable JSON-friendly scalars."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if pd.isna(value):
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return round(value, 6)
    if isinstance(value, (list, tuple, set)):
        return [_json_safe_cell(item) for item in value]
    if isinstance(value, dict):
        return {str(k): _json_safe_cell(v) for k, v in value.items()}
    return value


def _safe_rows(df: pd.DataFrame) -> list[list]:
    return [[_json_safe_cell(cell) for cell in row] for row in df.itertuples(index=False, name=None)]


def _safe_sheet_frame(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c) if str(c).strip() else f"column_{i+1}" for i, c in enumerate(df.columns)]
    for col in df.columns:
        df[col] = df[col].map(_json_safe_cell)
    return df


def _load_workbook_frames(file_path: str) -> dict[str, pd.DataFrame]:
    """Load sheets from either modern `.xlsx` or legacy `.xls` files."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        try:
            return pd.read_excel(file_path, sheet_name=None, engine="xlrd")
        except Exception:
            return {}

    return pd.read_excel(file_path, sheet_name=None, engine="openpyxl")


def _fallback_sheet_text(file_path: str, limit_sheets: int = 4, limit_rows: int = 40) -> dict[str, pd.DataFrame]:
    """Fallback for badly formatted workbooks: manual sheet scan via openpyxl."""
    try:
        import openpyxl
    except Exception:
        return {}

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return {}

    frames: dict[str, pd.DataFrame] = {}
    try:
        for sheet_index, sheet in enumerate(wb.worksheets):
            if sheet_index >= limit_sheets:
                break
            rows = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= limit_rows:
                    break
                rows.append([_json_safe_cell(cell) for cell in row])
            if not rows:
                continue
            headers = [f"column_{i+1}" for i in range(max(len(r) for r in rows))]
            frames[sheet.title] = pd.DataFrame(rows, columns=headers[: len(rows[0])])
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return frames


class ExcelExtractorTool:
    """
    Extracts tables, embedded images, formulas, and pivot table data from Excel files.
    Output strictly follows the NormalizedBlock contract in schemas.py.
    """
    name = "excel_extraction"

    def run(self, state: dict, config: dict) -> dict:
        file_path   = state["file_path"]
        document_id = state.get("document_id")
        blocks      = self._extract(file_path, document_id=document_id, config=config)
        state["blocks"] = as_dicts(blocks)
        state.setdefault("errors", [])
        return state

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _detect_language(self, text: str, default: str = "en") -> str:
        """Best-effort language code (ISO 639-1) for a block's text. Falls back to
        `default` when text is empty or langdetect isn't installed/confident."""
        if not text or not text.strip():
            return default
        try:
            from langdetect import detect
            return detect(text)
        except Exception:
            return default

    def _extract(
        self,
        file_path: str,
        document_id: Optional[str] = None,
        config: Optional[Dict[str, Any]] = None,
    ) -> List[NormalizedBlock]:
        blocks: List[NormalizedBlock] = []
        cfg      = config or {}
        doc_id   = str(document_id) if document_id else str(uuid.uuid4())
        filename = display_filename(file_path)

        try:
            wb = _load_workbook_frames(file_path)
        except Exception:
            wb = {}

        if not wb:
            wb = _fallback_sheet_text(file_path)

        if not wb:
            return blocks

        # ── 1. Table blocks ────────────────────────────────────────────
        for sheet_index, (sheet_name, df) in enumerate(wb.items()):
            try:
                df = df.dropna(how="all").dropna(axis=1, how="all")
                if df.empty:
                    continue
                df = _safe_sheet_frame(df).fillna("")
                safe_rows = _safe_rows(df)
                safe_headers = [str(c) for c in df.columns.tolist()]

                block_id = str(uuid.uuid4())
                cell_range = f"{sheet_name}!A1:{self._col_letter(len(df.columns))}{len(df) + 1}"
                try:
                    md_text = df.to_markdown(index=False)
                except Exception:
                    md_text = "\n".join(
                        " | ".join(map(str, row)) for row in ([safe_headers] + safe_rows[:20])
                    )

                blocks.append(NormalizedBlock(
                    block_id=block_id,
                    document_id=doc_id,
                    type="table",
                    text=md_text,
                    table_data={
                        "headers": safe_headers,
                        "rows": safe_rows,
                    },
                    source_ref=SourceRef(
                        filename=filename,
                        sheet=str(sheet_name),
                        sheet_index=sheet_index,
                    ),
                    confidence=cfg.get("extraction_confidence", 1.0),
                    language=self._detect_language(md_text, cfg.get("default_language", "en")),
                    metadata={
                        "cell_range":        cell_range,
                        "enrichment_failed": cfg.get("enrichment_failed_flag", False),
                    },
                ))

            except Exception:
                continue

        # ── 2. Embedded image blocks ───────────────────────────────────
        blocks.extend(self._extract_images(file_path, doc_id, filename, cfg))

        # ── 3. Formula blocks ──────────────────────────────────────────
        blocks.extend(self._extract_formulas(file_path, doc_id, filename, cfg))

        # ── 4. Pivot table blocks ──────────────────────────────────────
        blocks.extend(self._extract_pivots(file_path, doc_id, filename, cfg))

        return blocks

    def _extract_images(
        self,
        file_path: str,
        doc_id: str,
        filename: str,
        cfg: Dict[str, Any],
    ) -> List[NormalizedBlock]:
        import openpyxl

        blocks: List[NormalizedBlock] = []

        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception:
            return blocks

        out_dir = os.path.join("uploads", "images", doc_id)
        os.makedirs(out_dir, exist_ok=True)

        for sheet_index, sheet in enumerate(wb.worksheets):
            for image in getattr(sheet, "_images", []):
                try:
                    block_id   = str(uuid.uuid4())
                    image_data = None
                    if hasattr(image, "_data"):
                        try:
                            image_data = image._data()
                        except Exception:
                            image_data = None
                    if image_data is None and hasattr(image, "ref"):
                        image_data = image.ref.read() if hasattr(image.ref, "read") else bytes(image.ref)
                    if image_data is None:
                        continue
                    raw_path, preview_path = _save_image_blob(
                        image_data, out_dir, block_id, source_name=getattr(image, "path", None)
                    )

                    blocks.append(NormalizedBlock(
                        block_id=block_id,
                        document_id=doc_id,
                        type="image_caption",
                        text="",
                        source_ref=SourceRef(
                            filename=filename,
                            sheet=sheet.title,
                            sheet_index=sheet_index,
                        ),
                        confidence=cfg.get("extraction_confidence", 1.0),
                        language=cfg.get("default_language", "en"),
                        metadata={
                            "raw_image_path":    raw_path,
                            "image_path":        preview_path if preview_path else raw_path,
                            "pending_vision":    True,
                            "enrichment_failed": False,
                        },
                    ))

                except Exception:
                    continue

        return blocks

    def _extract_formulas(
        self,
        file_path: str,
        doc_id: str,
        filename: str,
        cfg: Dict[str, Any],
    ) -> List[NormalizedBlock]:
        import openpyxl

        blocks: List[NormalizedBlock] = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
        except Exception:
            return blocks

        for sheet_index, sheet in enumerate(wb.worksheets):
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type != "f" or cell.value is None:
                        continue
                    try:
                        formula  = str(cell.value)
                        cell_ref = f"{sheet.title}!{cell.coordinate}"
                        block_id = str(uuid.uuid4())

                        blocks.append(NormalizedBlock(
                            block_id=block_id,
                            document_id=doc_id,
                            type="text",
                            text=formula,
                            source_ref=SourceRef(
                                filename=filename,
                                sheet=sheet.title,
                                sheet_index=sheet_index,
                            ),
                            confidence=cfg.get("extraction_confidence", 1.0),
                            language=cfg.get("default_language", "en"),
                            metadata={
                                "cell_ref":          cell_ref,
                                "formula":           formula,
                                "sheet":             sheet.title,
                                "enrichment_failed": False,
                            },
                        ))

                    except Exception:
                        continue

        return blocks

    def _extract_pivots(
        self,
        file_path: str,
        doc_id: str,
        filename: str,
        cfg: Dict[str, Any],
    ) -> List[NormalizedBlock]:
        import openpyxl

        blocks: List[NormalizedBlock] = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception:
            return blocks

        for sheet_index, sheet in enumerate(wb.worksheets):
            for pivot in getattr(sheet, "_pivots", []):
                try:
                    pivot_name   = getattr(pivot, "name", "unnamed_pivot")
                    cache_def    = getattr(pivot, "cacheDefinition", None)
                    source_range = ""

                    if cache_def:
                        src = getattr(cache_def, "cacheSource", None)
                        if src and hasattr(src, "worksheetSource"):
                            source_range = getattr(src.worksheetSource, "ref", "")

                    headers = []
                    if cache_def and hasattr(cache_def, "cacheFields"):
                        headers = [
                            getattr(cf, "name", f"field_{i}")
                            for i, cf in enumerate(cache_def.cacheFields)
                        ]

                    rows = []
                    if cache_def and hasattr(cache_def, "records") and cache_def.records:
                        for record in cache_def.records.r:
                            rows.append([_json_safe_cell(getattr(item, "v", None)) for item in record])

                    if not headers and not rows:
                        continue

                    df       = pd.DataFrame(rows, columns=headers if headers else None)
                    block_id = str(uuid.uuid4())

                    blocks.append(NormalizedBlock(
                        block_id=block_id,
                        document_id=doc_id,
                        type="table",
                        text=df.to_markdown(index=False),
                        table_data={
                            "headers": headers,
                            "rows":    rows,
                        },
                        source_ref=SourceRef(
                            filename=filename,
                            sheet=sheet.title,
                            sheet_index=sheet_index,
                        ),
                        confidence=cfg.get("extraction_confidence", 1.0),
                        language=cfg.get("default_language", "en"),
                        metadata={
                            "pivot_name":         pivot_name,
                            "pivot_source_range": source_range,
                            "cell_range":         source_range,
                            "enrichment_failed":  False,
                        },
                    ))

                except Exception:
                    continue

        return blocks

    @staticmethod
    def _col_letter(n: int) -> str:
        result = ""
        while n:
            n, r = divmod(n - 1, 26)
            result = chr(65 + r) + result
        return result or "A"


# ------------------------------------------------------------------
# SANDBOX TEST
# ------------------------------------------------------------------
if __name__ == "__main__":
    test_file = "test-data/test.xlsx"

    mock_state  = {"file_path": test_file, "document_id": "doc-001"}
    mock_config = {"extraction_confidence": 0.95, "default_language": "en"}

    tool    = ExcelExtractorTool()
    state   = tool.run(mock_state, mock_config)
    results = state["blocks"]

    tables   = [b for b in results if b.type == "table" and "pivot_name" not in b.metadata]
    images   = [b for b in results if b.type == "image_caption"]
    formulas = [b for b in results if b.type == "text" and "formula" in b.metadata]
    pivots   = [b for b in results if b.type == "table" and "pivot_name" in b.metadata]

    print(f"Extracted {len(tables)} table(s), {len(images)} image(s), {len(formulas)} formula(s), {len(pivots)} pivot(s).\n")

    if tables:
        print("--- First table block ---")
        print(f"  text preview: {(tables[0].text or '')[:200]}")
        print(f"  headers: {tables[0].table_data['headers']}")

    if formulas:
        print("\n--- First formula block ---")
        print(f"  text: {formulas[0].text}")
        print(f"  cell_ref: {formulas[0].metadata['cell_ref']}")

    if images:
        print("\n--- First image block ---")
        print(f"  raw_image_path: {images[0].metadata['raw_image_path']}")
        print(f"  pending_vision: {images[0].metadata['pending_vision']}")

    if pivots:
        print("\n--- First pivot block ---")
        print(f"  pivot_name: {pivots[0].metadata['pivot_name']}")
        print(f"  text preview: {(pivots[0].text or '')[:200]}")
